from ctypes import alignment
from datetime import datetime
from msilib.schema import Font
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment
from rest_framework import generics
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import Product, Draft
from .serializers import ProductSerializer, DraftSerializer

class ProductListView(generics.ListAPIView):
    """
    Vista que maneja la solicitud GET para listar todos los productos.
    """
    def get_queryset(self):
        queryset = Product.objects.all()
        checkedParam = self.request.query_params.get('checked', None)
        if checkedParam:
            queryset = queryset.filter(checked = checkedParam)
        return queryset
    
    serializer_class = ProductSerializer

class UpdateStockView(APIView):
    """
    Vista que maneja la solicitud PUT para actualizar el stock de un producto específico y su estado "checked".
    """
    def put(self, request, product_id):
        product = generics.get_object_or_404(Product, id=product_id)
        new_stock = request.data.get('stock')
        new_checked = request.data.get('checked', None)  # Obtener el estado "checked" desde la solicitud

        if new_stock is not None:
            try:
                product.stock = int(new_stock)
                if new_checked is not None:
                    product.checked = bool(new_checked)  # Actualizar el estado "checked" del producto
                product.save()
                return Response({
                    'message': 'Stock and checked status updated successfully',
                    'stock': product.stock,
                    'checked': product.checked
                })
            except ValueError:
                return Response({'error': 'Invalid stock value'}, status=400)
        return Response({'error': 'Stock value is required'}, status=400)
    

class CreateProductView(generics.CreateAPIView):
    """
    Vista que maneja la solicitud POST para crear un producto.
    """
    queryset = Product.objects.all()
    serializer_class = ProductSerializer


class DraftView(generics.ListCreateAPIView):
    """
    Vista que maneja las solicitudes GET para listar todos los drafts
    y POST para crear un nuevo draft.
    """
    queryset = Draft.objects.all()
    serializer_class = DraftSerializer

    def get(self, request, *args, **kwargs):
        """
        Maneja la solicitud GET para obtener todos los drafts.
        """
        drafts = self.get_queryset()
        serializer = self.get_serializer(drafts, many=True)
        return Response(serializer.data)

    def post(self, request, *args, **kwargs):
        """
        Maneja la solicitud POST para crear un nuevo draft.
        """
        return super().post(request, *args, **kwargs)

class ProductExcelReportView(APIView):
    """
    Vista que maneja la generación y descarga de un archivo Excel con los datos de todos los productos.
    """
    def get(self, request, *args, **kwargs):
        # Extraer los parámetros de la solicitud desde el frontend
        inventory_name = request.GET.get('inventory_name', 'Inventario Desconocido')
        created_by = request.GET.get('created_by', 'Desconocido')
        created_at = request.GET.get('created_at', str(datetime.now().strftime('%Y-%m-%d')))

        # Crear un libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inventario de Productos"

        # Estilo para los encabezados
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")

        # Encabezado del reporte
        sheet.merge_cells('A1:F1')
        report_title = sheet.cell(row=1, column=1)
        report_title.value = f"Inventario: {inventory_name}"
        report_title.font = header_font
        report_title.alignment = header_alignment

        sheet.merge_cells('A2:F2')
        report_subtitle = sheet.cell(row=2, column=1)
        report_subtitle.value = f"Creado por: {created_by} | Fecha de creación: {created_at}"
        report_subtitle.font = header_font
        report_subtitle.alignment = header_alignment

        # Encabezados de columnas
        headers = ["ID", "Nombre", "Descripción", "Stock", "Precio", "Estado"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col_num)  # Cambiado de 1 a 4 para dejar espacio al encabezado
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment

        # Obtener todos los productos y agregar filas
        products = Product.objects.all()
        for row_num, product in enumerate(products, 5):  # Cambiado de 2 a 5 para dejar espacio al encabezado
            sheet.cell(row=row_num, column=1).value = product.id
            sheet.cell(row=row_num, column=2).value = product.name
            sheet.cell(row=row_num, column=3).value = product.description
            sheet.cell(row=row_num, column=4).value = product.stock
            sheet.cell(row=row_num, column=5).value = float(product.price)  # Convertir a float
            sheet.cell(row=row_num, column=6).value = "Checkeado" if product.checked else "No checkeado"

        # Ajustar el ancho de las columnas
        column_widths = {
            'A': 0,
            'B': 0,
            'C': 0,
            'D': 0,
            'E': 0,
            'F': 0,
        }

        for col_letter in column_widths:
            for cell in sheet[col_letter]:
                if not isinstance(cell, openpyxl.cell.MergedCell):
                    column_widths[col_letter] = max(column_widths[col_letter], len(str(cell.value)))

        for col_letter, col_width in column_widths.items():
            sheet.column_dimensions[col_letter].width = col_width + 2

        # Preparar la respuesta HTTP para la descarga del archivo
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="productos_inventario.xlsx"'
        workbook.save(response)
        return response